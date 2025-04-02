

import openpyxl

class SearchemployeeData:
    test_search_Data = [{"empname":"Gowthami"}]


class ExcelWriter:
    @staticmethod
    def write_employee_data(sheet_name,employee_data):
        file_path = "C:\\Users\\Dell\\Downloads\\reademployee.xlsx"
        sheet_name = "WriteData"
        """Write employee details to an Excel sheet."""
        # try:
        #     workbook = openpyxl.load_workbook(file_path)
        # except FileNotFoundError:
        #     workbook = openpyxl.Workbook()
        #
        # if sheet_name in workbook.sheetnames:
        #     sheet = workbook[sheet_name]
        #     sheet = workbook.create_sheet(sheet_name)
        #     #sheet.append(["Employee ID", "Employee Name", "Department", "Designation", "Email", "Mobile"])
        #
        # else:
        #     raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")
        book = openpyxl.load_workbook("C:\\Users\\Dell\\Downloads\\reademployee.xls")

        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")


        for i in range(1,6):
            for c in range(1,4):
                sheet.cell(row=i,column=c).value = "empname"



            # sheet.append([
            #     employee_data["Employee ID"],
            #     employee_data["Employee Name"],
            #     employee_data["Department"],
            #     employee_data["Designation"],
            #     employee_data["Email"],
            #     employee_data["Mobile"]
            # ])

        book.save(file_path)



