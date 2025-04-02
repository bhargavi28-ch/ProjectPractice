import openpyxl



class Excelutils:
    def getexcel(self,driver):
        self.driver = driver


    @staticmethod
    def getRowcount(file, sheetName):
        Workbook = openpyxl.load_workbook(file,data_only=True)
        sheet = Workbook[sheetName]
        row_count = sheet.max_row
        Workbook.close()
        return row_count

    @staticmethod
    def getColoumncount(file, sheetName):
        Workbook = openpyxl.load_workbook(file,data_only=True)
        sheet = Workbook[sheetName]
        col_count = sheet.max_column
        Workbook.close()
        return col_count

    @staticmethod
    def readData(file, sheetName, rownum, coloumnno):
        Workbook = openpyxl.load_workbook(file,data_only=True)
        sheet = Workbook[sheetName]
        value = sheet.cell(row=rownum,column=coloumnno).value
        Workbook.close()
        return value



    @staticmethod
    def writeData(file, sheetName , rownum, coloumnno, data):
        Workbook = openpyxl.load_workbook(file)
        sheet = Workbook[sheetName]
        # sheet.cell(row=rownum,column=coloumnno).value = data
        sheet.cell(row=rownum,column=coloumnno,value = data)
        Workbook.save(file)
        Workbook.close()

        Workbook = openpyxl.load_workbook(file)
        Workbook.save(file)
        Workbook.close()
