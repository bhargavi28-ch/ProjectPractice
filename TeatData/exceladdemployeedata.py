import openpyxl
# book = openpyxl.load_workbook("C:\\Users\\Dell\\OneDrive - paccore\\Desktop\\addemployee.xlsx")
# sheet = book.active
#
# Dict = {}
#
# for i in range(1,sheet.max_row):
#         for j in range(1,sheet.max_column):
#             Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
#
#
# print(Dict)



@staticmethod
def getTestData():
    employee_list = []
    book = openpyxl.load_workbook("C:\\Users\\Dell\\OneDrive - paccore\\Desktop\\twoemployees.xlsx")
    sheet = book.active

    for i in range(1, sheet.max_row + 1, 2):
        emp1 = {}
        emp2 = {}

        for j in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=j).value
            emp1[header] = sheet.cell(row=i, column=j).value
            if i + 1 <= sheet.max_row:
                emp2[header] = sheet.cell(row=i+1, column=j).value
        employee_list.append(emp1)
        employee_list.append(emp2)


    print(employee_list)
    return employee_list


import openpyxl


class AddemployepageData:

    @staticmethod
    def getTestData(sheet_name="Sheet1"):  # Default is Sheet1, can be changed to Sheet2 or any other
        employee_list = []

        # Load the workbook
        book = openpyxl.load_workbook("C:\\Users\\Dell\\OneDrive - paccore\\Desktop\\addemployee.xlsx")

        # Check if the given sheet_name exists
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

        # Start iterating through the rows, in steps of 2 to pair employee data
        for i in range(2, sheet.max_row + 1, 2):  # Start from row 2 and iterate in steps of 2
            emp1 = {}  # Dictionary to store first employee data
            emp2 = {}  # Dictionary to store second employee data

            # Loop through all columns for each row (for both employees)
            for j in range(1, sheet.max_column + 1):  # Loop through columns
                header = sheet.cell(row=1, column=j).value  # Get the header (column name)

                # Store data for employee 1 (row i)
                emp1[header] = sheet.cell(row=i, column=j).value

                # Check if there's a second employee (row i+1) and store the data
                if i + 1 <= sheet.max_row:
                    emp2[header] = sheet.cell(row=i + 1, column=j).value

            # Add both employees' data to the employee list
            employee_list.append((emp1, emp2))

        return employee_list






