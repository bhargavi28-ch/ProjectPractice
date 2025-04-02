import openpyxl


class CreatecompanyData:


    test_create_company_Data = [{"Company_Name":"ABC Pvt Ltd",
                                 "Short_Name":"ABC",
                                 "Company_Type":"Private Limited",
                                 "Company_Registration_No.":"123456789",
                                 "Date_of_Registration":"10-02-2022",
                                 "Registration_Type":"LLC",
                                 "Nature_of_Business":"IT Services",
                                 "Allowed_Leaves_Per_Month":"2",
                                 "select_company":"ABC Pvt Ltd"}]
    upload_file = "C:/Users/Dell/Downloads/blue_profile_icon.png"




    # @staticmethod
    # def getTestData():
    #     Dict = { }
    #     book = openpyxl.load_workbook("C:\\Users\\Dell\\Downloads\\companycreation.xlsx")
    #     sheet = book.active
    #
    #     for i in range(1, sheet.max_row):
    #         for j in range(1, sheet.max_column):
    #             Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    #             print(sheet.cell(row=i, column=j).value)
    #
    #     return[Dict]
