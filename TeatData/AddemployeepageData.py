import openpyxl


class AddemployepageData:

    test_addemployeepage_Data = [{"firstname":"pranathi",
                                  "lastname":"mounika",
                                  "fathername":"venkatesh",
                                  "mobilenumber":"9876545678",
                                  "companyemail":"bhargavi.ch+134@paccore.com",
                                  "personalemail":"bhargavich.207+144@gmail.com",
                                  "dob":"30-05-2000"
                                  }]

    #for on employee cretaed from excel sheet
    # @staticmethod
    # def getTestData():
    #     Dict = {}
    #     book = openpyxl.load_workbook("C:\\Users\\Dell\\Downloads\\Oneemployee.xlsx")
    #     sheet = book.active
    #
    #
    #
    #     for i in range(2, sheet.max_row):
    #         for j in range(2, sheet.max_column):
    #             Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    #
    #     return[Dict]

    #for 2 employees
    # @staticmethod
    # def getTestData():
    #     employee_list = []
    #     book = openpyxl.load_workbook("C:\\Users\\Dell\\OneDrive - paccore\\Desktop\\twoemployees.xlsx")
    #     sheet = book.active
    #
    #
    #     for i in range(2, sheet.max_row+1, 2):
    #         emp1 = {}
    #         emp2 = {}
    #
    #         for j in range(1, sheet.max_column+1):
    #             header = sheet.cell(row=1, column=j).value
    #             emp1[header] = sheet.cell(row=i, column=j).value
    #             if i+1 <= sheet.max_row:
    #                 emp2[header] = sheet.cell(row=i+1, column=j).value
    #
    #         employee_list.append(emp1)
    #         employee_list.append(emp2)
    #     return employee_list
    #
    #excel form is one but import another sheet
    @staticmethod
    def getTestData(sheet_name = "ReadData"):
        employee_list = []
        book = openpyxl.load_workbook("C:\\Users\\Dell\\Downloads\\reademployee.xlsx")

        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

        for i in range(2, sheet.max_row+1,10):

            emp1 = {}
            emp2 = {}

            for j in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=j).value
                emp1[header] = sheet.cell(row=i, column=j).value
                if i+1 <= sheet.max_row:
                    emp2[header] = sheet.cell(row=i+1, column=j).value

            employee_list.append(emp1)
            employee_list.append(emp2)



        return employee_list

